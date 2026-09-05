"""Sanare 12-month budget model (Aug 2026 -> Jul 2027).

Single source of truth for the plan numbers. Generates ``budget_12mo.xlsx`` and
exposes ``build()`` for ``budget_sync.py`` to upload into Odoo.

All inputs are USD; Odoo stores THB at ``FX`` (34 THB/USD).
"""
import datetime

FX = 34.0                       # THB per USD (estimate)
FUNDING_USD = 350_000.0         # capital received at period start
COGS_RATIO = 0.50               # 50% gross margin

# --- 4 products: name, USD per cm2, share of volume ---
PRODUCTS = [
    ("dHACM Amniotic Membrane Allograft (dual-layer)", 200.0, 0.45),
    ("Amniotic Membrane Allograft (single-layer)", 140.0, 0.30),
    ("Collagen Dressing", 70.0, 0.15),
    ("Antimicrobial Dressing", 90.0, 0.10),
]
BLENDED_ASP = sum(p * s for _, p, s in PRODUCTS)  # 151.5 USD/cm2

# --- 12 monthly windows Aug-2026 .. Jul-2027 ---
# cm2 sold per month (0 pre-launch; sales kick off Feb-2027)
VOLUME_CM2 = [0, 0, 0, 0, 0, 0, 5_000, 10_000, 18_000, 28_000, 38_000, 50_000]

# --- monthly opex (USD) per EOS category ---
OPEX_USD = {
    "payroll":            [20_000, 20_000, 22_000, 24_000, 26_000, 28_000, 30_000, 33_000, 36_000, 40_000, 43_000, 45_000],
    "regulatory":         [20_000, 20_000, 10_000, 5_000, 5_000, 5_000, 5_000, 5_000, 5_000, 5_000, 5_000, 5_000],
    "legal_prof":         [8_000] * 12,
    "sales_marketing":    [3_000, 3_000, 4_000, 5_000, 8_000, 12_000, 18_000, 20_000, 22_000, 24_000, 26_000, 28_000],
    "travel":             [2_000, 2_000, 3_000, 3_000, 4_000, 5_000, 6_000, 7_000, 8_000, 9_000, 10_000, 10_000],
    "office_it_ga":       [6_000] * 12,
    "inventory_purchases":[0, 0, 0, 0, 0, 100_000, 60_000, 40_000, 30_000, 20_000, 10_000, 0],
    "other_capex":        [0, 40_000, 10_000, 5_000, 0, 0, 5_000, 0, 0, 0, 5_000, 0],
}

CATEGORIES = [
    "revenue", "cogs", "payroll", "regulatory", "legal_prof", "sales_marketing",
    "travel", "office_it_ga", "inventory_purchases", "other_capex",
    "other_opex", "other_cash",
]

CATEGORY_LABELS = {
    "revenue": "Revenue", "cogs": "COGS", "payroll": "Payroll & Benefits",
    "regulatory": "Regulatory", "legal_prof": "Legal & Professional",
    "sales_marketing": "Sales & Marketing", "travel": "Travel",
    "office_it_ga": "Office / IT / G&A", "inventory_purchases": "Inventory Purchases",
    "other_capex": "Other CapEx", "other_opex": "Other Operating", "other_cash": "Other Cash Uses",
}


def month_windows():
    """12 (start, end) date tuples Aug-2026 .. Jul-2027."""
    windows = []
    y, m = 2026, 8
    for _ in range(12):
        start = datetime.date(y, m, 1)
        if m == 12:
            nxt = datetime.date(y + 1, 1, 1)
        else:
            nxt = datetime.date(y, m + 1, 1)
        windows.append((start, nxt - datetime.timedelta(days=1)))
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return windows


def build():
    """Return {months, categories, labels, usd, thb, products, funding_thb}."""
    months = [d.strftime("%Y-%m") for d, _ in month_windows()]

    usd = {c: [0.0] * 12 for c in CATEGORIES}
    for i, vol in enumerate(VOLUME_CM2):
        usd["revenue"][i] = vol * BLENDED_ASP
        usd["cogs"][i] = vol * BLENDED_ASP * COGS_RATIO
    for cat, amounts in OPEX_USD.items():
        usd[cat] = list(amounts)

    thb = {c: [round(v * FX) for v in usd[c]] for c in CATEGORIES}

    products = [
        {
            "name": name,
            "usd_price": price,
            "share": share,
            "thb_price": round(price * FX),
            "thb_cost": round(price * FX * COGS_RATIO),
        }
        for name, price, share in PRODUCTS
    ]

    return {
        "months": months,
        "categories": CATEGORIES,
        "labels": CATEGORY_LABELS,
        "usd": usd,
        "thb": thb,
        "products": products,
        "funding_usd": FUNDING_USD,
        "funding_thb": round(FUNDING_USD * FX),
        "blended_asp": BLENDED_ASP,
    }


def write_xlsx(path="budget_12mo.xlsx"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    d = build()
    wb = openpyxl.Workbook()

    # --- Budget sheet (THB) ---
    ws = wb.active
    ws.title = "Budget (THB)"
    header = ["Category"] + d["months"]
    ws.append(header)
    for c in d["categories"]:
        ws.append([d["labels"][c]] + d["thb"][c])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    # --- Budget sheet (USD) ---
    ws2 = wb.create_sheet("Budget (USD)")
    ws2.append(header)
    for c in d["categories"]:
        ws2.append([d["labels"][c]] + d["usd"][c])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    # --- Products ---
    ws3 = wb.create_sheet("Products")
    ws3.append(["Product", "USD/cm2", "Share", "THB/cm2", "THB cost/cm2"])
    for p in d["products"]:
        ws3.append([p["name"], p["usd_price"], p["share"], p["thb_price"], p["thb_cost"]])

    # --- Summary ---
    ws4 = wb.create_sheet("Summary")
    ws4.append(["Metric", "Value"])
    ws4.append(["Blended ASP (USD/cm2)", d["blended_asp"]])
    ws4.append(["Funding (USD)", d["funding_usd"]])
    ws4.append(["Funding (THB)", d["funding_thb"]])
    ws4.append(["End-rate volume (cm2/mo)", VOLUME_CM2[-1]])
    ws4.append(["End-rate revenue (THB/mo)", round(VOLUME_CM2[-1] * BLENDED_ASP * FX)])
    ws4.append(["Total 12-mo revenue (USD)", round(sum(d["usd"]["revenue"]))])
    ws4.append(["FX (THB/USD)", FX])

    wb.save(path)
    return path


if __name__ == "__main__":
    d = build()
    path = write_xlsx()
    print(f"Wrote {path}")
    print(f"Blended ASP = ${d['blended_asp']:.1f}/cm2 | end-rate = {VOLUME_CM2[-1]:,} cm2/mo")
    print(f"Total 12-mo revenue = ${sum(d['usd']['revenue']):,.0f} | funding ${d['funding_usd']:,.0f}")
    for c in d["categories"]:
        tot = sum(d["thb"][c])
        if tot:
            print(f"  {d['labels'][c]:22s} total {tot:>15,} THB")
