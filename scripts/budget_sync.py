"""Upload the budget model into Odoo and download budget + actuals to Excel.

Upload   : python3 scripts/budget_sync.py upload
Download : python3 scripts/budget_sync.py download [out.xlsx]

Budget is two-way (model -> budget.lines -> Excel). Actuals are one-way
(Odoo eos.financial.period -> Excel).
"""
import sys

from sanare_odoo import call
from budget_model import CATEGORIES, CATEGORY_LABELS, build, month_windows

COMPANY_ID = 1
BUDGET_NAME = "Sanare FY27 Budget (Aug-26 to Jul-27)"


def _post_by_category():
    posts = call("account.budget.post", "search_read",
                 [[("company_id", "=", COMPANY_ID), ("eos_category", "!=", False)]],
                 {"fields": ["id", "eos_category"]})
    return {p["eos_category"]: p["id"] for p in posts}


def upload():
    d = build()
    posts = _post_by_category()
    missing = [c for c in CATEGORIES if c not in posts]
    if missing:
        raise SystemExit("Missing budget posts for %s — run eos_financials_setup.py first" % missing)

    windows = month_windows()
    found = call("budget.budget", "search_read", [[("name", "=", BUDGET_NAME)]], {"fields": ["id"]})
    if found:
        bid = found[0]["id"]
    else:
        bid = call("budget.budget", "create", [{
            "name": BUDGET_NAME,
            "date_from": windows[0][0].isoformat(),
            "date_to": windows[-1][1].isoformat(),
            "company_id": COMPANY_ID,
        }])

    # wipe + rewrite lines for this budget (idempotent)
    existing = call("budget.lines", "search", [[("budget_id", "=", bid)]])
    if existing:
        call("budget.lines", "unlink", [existing])

    created = 0
    for cat in CATEGORIES:
        post_id = posts[cat]
        for mi, (start, end) in enumerate(windows):
            thb = d["thb"][cat][mi]
            if thb == 0:
                continue
            call("budget.lines", "create", [{
                "budget_id": bid,
                "general_budget_id": post_id,
                "date_from": start.isoformat(),
                "date_to": end.isoformat(),
                "planned_amount": float(thb),
            }])
            created += 1

    periods = call("eos.financial.period", "search", [[]])
    if periods:
        call("eos.financial.period", "action_refresh", [periods])

    print(f"Uploaded {created} budget.lines into budget #{bid} ({BUDGET_NAME})")
    print(f"Refreshed {len(periods)} financial periods")


def download(path="budget_actuals.xlsx"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    d = build()
    months = d["months"]
    mi_of = {m: i for i, m in enumerate(months)}

    posts = _post_by_category()
    cat_by_post = {v: k for k, v in posts.items()}

    # --- budget matrix from budget.lines ---
    lines = call("budget.lines", "search_read",
                 [[("general_budget_id", "in", list(posts.values()))]],
                 {"fields": ["general_budget_id", "date_from", "planned_amount"]})
    budget = {c: [0.0] * 12 for c in CATEGORIES}
    for l in lines:
        cat = cat_by_post[l["general_budget_id"][0]]
        mi = mi_of.get(l["date_from"][:7])
        if mi is not None:
            budget[cat][mi] = l["planned_amount"]

    # --- actuals from eos.financial.period ---
    fields = ["date_from", "ending_cash", "net_cash_burn", "runway_months"] + CATEGORIES
    periods = call("eos.financial.period", "search_read", [[]],
                   {"fields": fields, "order": "date_from"})
    actuals = {c: [0.0] * 12 for c in CATEGORIES}
    cash = {"month": [], "ending_cash": [], "net_cash_burn": [], "runway_months": []}
    for p in periods:
        mi = mi_of.get((p["date_from"] or "")[:7])
        if mi is None:
            continue
        for c in CATEGORIES:
            actuals[c][mi] = p.get(c, 0.0) or 0.0
        cash["month"].append(months[mi])
        cash["ending_cash"].append(p.get("ending_cash", 0.0) or 0.0)
        cash["net_cash_burn"].append(p.get("net_cash_burn", 0.0) or 0.0)
        cash["runway_months"].append(p.get("runway_months", 0.0) or 0.0)

    wb = openpyxl.Workbook()

    def matrix_sheet(title, data):
        ws = wb.create_sheet(title)
        ws.append(["Category"] + months)
        for c in CATEGORIES:
            ws.append([CATEGORY_LABELS[c]] + data[c])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
        return ws

    matrix_sheet("Budget (THB)", budget)
    matrix_sheet("Actuals (THB)", actuals)

    ws = wb.create_sheet("Cash & Runway")
    ws.append(cash["month"])
    ws.append(["Ending Cash (THB)"] + cash["ending_cash"])
    ws.append(["Net Cash Burn (THB)"] + cash["net_cash_burn"])
    ws.append(["Runway (months)"] + cash["runway_months"])

    wb.save(path)
    print(f"Wrote {path}")
    return path


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "download"
    if cmd == "upload":
        upload()
    elif cmd == "download":
        download(sys.argv[2] if len(sys.argv) > 2 else "budget_actuals.xlsx")
    else:
        raise SystemExit("usage: budget_sync.py [upload|download [out.xlsx]]")
