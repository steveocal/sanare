#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parity check: ``readiness_engine`` vs the real workbook.

Loads ``Sanare_Management_Reporting_Engine_v2_ORDERED.xlsx``, feeds the
``Task Tracker`` rows through the ported Python functions and asserts that every
figure matches the workbook's own cached cell values on both
``'Launch Readiness'`` and ``'Report 03 Thailand'``.

Usage::

    python3 eos_dashboard/scripts/verify_vs_spreadsheet.py \
        [--xlsx PATH] [--asof YYYY-MM-DD] [-v]

``--asof`` stands in for the spreadsheet's ``TODAY()`` in the Health column and
defaults to today. The workbook was last recalculated in the Sep-Oct 2026
window, so the cached Health values reproduce for any ``--asof`` in that range.
Readiness %, the weighted total and the Red/Yellow counts are date-independent.

Exit code is 0 on full parity, 1 otherwise.
"""

from __future__ import annotations

import argparse
import datetime
import os
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_HERE, os.pardir, "models"))

import readiness_engine as eng  # noqa: E402

try:
    import openpyxl  # noqa: E402
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required: pip install openpyxl")

DEFAULT_XLSX = os.path.join(
    _HERE, os.pardir, os.pardir, "Sanare_Management_Reporting_Engine_v2_ORDERED.xlsx"
)

# 'Launch Readiness' column indexes (1-based) and the Thailand block row range.
LR_SHEET = "Launch Readiness"
LR_FIRST_ROW, LR_LAST_ROW = 4, 12          # rows 4..12 = the 9 Thailand workstreams
LR_COL_MARKET, LR_COL_WEIGHT = 1, 3        # A, C
LR_COL_READINESS, LR_COL_STATUS = 4, 5     # D, E
LR_SUMMARY_ROW = 24                        # B24/C24/D24 = Thailand summary

# 'Report 03 Thailand' cells that mirror the Launch Readiness block.
R3_SHEET = "Report 03 Thailand"
R3_FIRST_ROW = 8                           # C8..C16 readiness, D8..D16 health
R3_COL_READINESS, R3_COL_STATUS = 3, 4     # C, D
R3_OVERALL, R3_RED, R3_YELLOW = "B5", "D5", "F5"

# 'Task Tracker' column indexes (1-based).
TT_SHEET = "Task Tracker"
TT_COL_ROCK, TT_COL_PRIORITY = 4, 8        # D, H
TT_COL_DUE, TT_COL_STATUS = 10, 11         # J, K
TT_COL_CRIT_PATH = 14                      # N

TOL = 1e-6


def load_task_rows(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        rock = ws.cell(r, TT_COL_ROCK).value
        if rock in (None, ""):
            continue
        rows.append(eng.TaskRow(
            rock=str(rock),
            status=ws.cell(r, TT_COL_STATUS).value or "",
            due=ws.cell(r, TT_COL_DUE).value,
            priority=ws.cell(r, TT_COL_PRIORITY).value or "",
            critical_path=ws.cell(r, TT_COL_CRIT_PATH).value or "",
        ))
    return rows


class Checker:
    def __init__(self, verbose=False):
        self.verbose = verbose
        self.failures = 0
        self.checks = 0

    def _fmt(self, v):
        if isinstance(v, float):
            return "%.10g" % v
        return str(v)

    def eq(self, label, got, want, tol=0.0):
        self.checks += 1
        if isinstance(want, (int, float)) and isinstance(got, (int, float)):
            ok = abs(float(got) - float(want)) <= (tol or 0.0)
        else:
            ok = got == want
        mark = "ok  " if ok else "FAIL"
        if not ok:
            self.failures += 1
        if self.verbose or not ok:
            print(f"  [{mark}] {label:<46} got={self._fmt(got):<16} want={self._fmt(want)}")
        return ok


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--asof", default=None,
                    help="YYYY-MM-DD used for the Health column (default: today)")
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args(argv)

    as_of = (datetime.date.fromisoformat(args.asof) if args.asof
             else datetime.date.today())

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    tasks = load_task_rows(wb[TT_SHEET])
    result = eng.compute_thailand_readiness(tasks, today=as_of)
    engine_rows = result["rows"]

    print(f"xlsx      : {os.path.relpath(args.xlsx)}")
    print(f"as-of     : {as_of.isoformat()}  (Health column only)")
    print(f"task rows : {len(tasks)} with a Rock ID")
    print()

    ck = Checker(verbose=args.verbose)

    # ---- 'Launch Readiness' Thailand block -------------------------------
    lr = wb[LR_SHEET]
    print(f"[{LR_SHEET}] rows {LR_FIRST_ROW}-{LR_LAST_ROW} + summary row {LR_SUMMARY_ROW}")
    for i, row_no in enumerate(range(LR_FIRST_ROW, LR_LAST_ROW + 1)):
        er = engine_rows[i]
        assert lr.cell(row_no, LR_COL_MARKET).value == "Thailand", row_no
        want_readiness = float(lr.cell(row_no, LR_COL_READINESS).value or 0.0)
        want_status = lr.cell(row_no, LR_COL_STATUS).value
        want_weight = float(lr.cell(row_no, LR_COL_WEIGHT).value or 0.0)
        ck.eq(f"{er['label']} weight", er["weight"], want_weight, TOL)
        ck.eq(f"{er['label']} readiness", er["readiness"], want_readiness, TOL)
        ck.eq(f"{er['label']} status", er["status"], want_status)

    want_overall = float(lr.cell(LR_SUMMARY_ROW, 2).value or 0.0)   # B24
    want_red = int(lr.cell(LR_SUMMARY_ROW, 3).value or 0)           # C24
    want_yellow = int(lr.cell(LR_SUMMARY_ROW, 4).value or 0)        # D24
    ck.eq("overall weighted readiness (B24)", result["overall_readiness"], want_overall, TOL)
    ck.eq("red workstreams (C24)", result["red_count"], want_red)
    ck.eq("yellow workstreams (D24)", result["yellow_count"], want_yellow)

    # ---- 'Report 03 Thailand' mirror cells -----------------------------
    r3 = wb[R3_SHEET]
    print(f"\n[{R3_SHEET}] rows {R3_FIRST_ROW}-{R3_FIRST_ROW + 8} + {R3_OVERALL}/{R3_RED}/{R3_YELLOW}")
    for i in range(9):
        row_no = R3_FIRST_ROW + i
        er = engine_rows[i]
        want_readiness = float(r3.cell(row_no, R3_COL_READINESS).value or 0.0)
        want_status = r3.cell(row_no, R3_COL_STATUS).value
        ck.eq(f"{er['label']} readiness (C{row_no})", er["readiness"], want_readiness, TOL)
        ck.eq(f"{er['label']} health (D{row_no})", er["status"], want_status)
    ck.eq(f"overall ({R3_OVERALL})", result["overall_readiness"], float(r3[R3_OVERALL].value or 0.0), TOL)
    ck.eq(f"red ({R3_RED})", result["red_count"], int(r3[R3_RED].value or 0))
    ck.eq(f"yellow ({R3_YELLOW})", result["yellow_count"], int(r3[R3_YELLOW].value or 0))

    print()
    print("-" * 64)
    if ck.failures:
        print(f"PARITY FAILED: {ck.failures} of {ck.checks} checks mismatched.")
        return 1
    print(f"PARITY OK: all {ck.checks} checks match the workbook's cached values.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
